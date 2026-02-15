# src/storage/excel_storage.py
from datetime import datetime
from pathlib import Path
from typing import Any, Callable
import pandas as pd
from openpyxl import load_workbook

from src.storage.base import BaseStorage


class ExcelStorage(BaseStorage):

    def __init__(self, filepath: Path, schema_func: Callable[[], pd.DataFrame]):
        self.filepath = filepath
        self.schema_func = schema_func
        self.filepath.parent.mkdir(parents=True, exist_ok=True)

    def create_sheet(self) -> pd.DataFrame:
        schema = self.schema_func()
        schema.to_excel(self.filepath, index=False)
        return schema

    def load_sheet(self) -> pd.DataFrame:
        return pd.read_excel(self.filepath)

    def register_data(self, new_row: dict[str, Any]) -> None:
        if not self.filepath.exists():
            self.create_sheet()

        wb = load_workbook(self.filepath)
        ws = wb.active

        ws.append(list(new_row.values()))

        wb.save(self.filepath)

    def get_last_row(self, before_date: datetime) -> pd.Series | None:
        if not self.filepath.exists():
            return None

        df = self.load_sheet()
        if df.empty:
            return None

        df["date"] = pd.to_datetime(df["date"])
        df = df[df["date"] < before_date]
        if df.empty:
            return None

        return df.sort_values("date").iloc[-1]

    def get_data(self, year: int | None = None, month: int | None = None, filepath: Path | None = None) -> pd.DataFrame:
        target_filepath = filepath if filepath else self.filepath

        if not target_filepath.exists():
            return self.schema_func()

        df = pd.read_excel(target_filepath)
        df["date"] = pd.to_datetime(df["date"])

        if year is not None:
            df = df[df["date"].dt.year == year]
        if month is not None:
            df = df[df["date"].dt.month == month]

        return df

    def order_by_date(self, df: pd.DataFrame) -> pd.DataFrame:
        df["date"] = pd.to_datetime(df["date"])
        return df.sort_values("date")