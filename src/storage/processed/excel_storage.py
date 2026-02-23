from datetime import datetime
from pathlib import Path
from typing import Any, Callable
import pandas as pd
from openpyxl import load_workbook

from src.storage.processed import BaseProcessedStorage


class ExcelProcessedStorage(BaseProcessedStorage):
    """Implementação de armazenamento de dados processados usando arquivos Excel, com métodos para criação, carregamento, registro e consulta de dados organizados por data.

    Attributes:
        filepath (Path): Caminho para o arquivo Excel onde os dados processados serão armazenados.
        schema_func (Callable[[], pd.DataFrame]): Função que retorna um DataFrame vazio com a estrutura (colunas) apropriada para os dados processados, usada para criar a planilha inicial quando o arquivo não existir.
        filepath.parent.mkdir(parents=True, exist_ok=True): Garante que o diretório para o arquivo Excel exista, criando-o se necessário.
    """

    def __init__(self, filepath: Path, schema_func: Callable[[], pd.DataFrame]) -> None:
        """Inicializa o armazenamento de dados processados em Excel, definindo o caminho do arquivo e a função de esquema para criar a planilha inicial.

        Args:
            filepath (Path): Caminho para o arquivo Excel onde os dados processados serão armazenados.
            schema_func (Callable[[], pd.DataFrame]): Função que retorna um DataFrame vazio com a estrutura (colunas) apropriada para os dados processados, usada para criar a planilha inicial quando o arquivo não existir.
        """
        self.filepath = filepath
        self.schema_func = schema_func
        self.filepath.parent.mkdir(parents=True, exist_ok=True)

    def create_sheet(self) -> pd.DataFrame:
        """Cria uma nova planilha Excel com a estrutura definida pela função de esquema, salvando-a no caminho especificado.
        Retorna o DataFrame criado.

        Returns:
            pd.DataFrame: DataFrame vazio com a estrutura definida pela função de esquema, salvo no arquivo Excel especificado.
        """
        schema = self.schema_func()
        schema.to_excel(self.filepath, index=False)
        return schema

    def load_sheet(self) -> pd.DataFrame:
        """Carrega a planilha Excel existente com os dados processados.
        Se o arquivo não existir, retorna um DataFrame vazio com a estrutura definida pela função de esquema.

        Returns:
            pd.DataFrame: DataFrame contendo os dados processados atualmente armazenados.
        """
        return pd.read_excel(self.filepath)

    def register_data(self, new_row: dict[str, Any]) -> None:
        """Registra uma nova linha de dados processados na planilha Excel, garantindo que os dados sejam organizados por data e que a planilha seja atualizada corretamente.
        Se o arquivo não existir, cria uma nova planilha usando a função de esquema.

        Args:
            new_row (dict[str, Any]): Dicionário contendo os dados a serem registrados, onde as chaves correspondem às colunas da planilha Excel.
        """
        if not self.filepath.exists():
            self.create_sheet()

        wb = load_workbook(self.filepath)
        ws = wb.active

        ws.append(list(new_row.values()))

        wb.save(self.filepath)

    def get_last_row(self, before_date: datetime) -> pd.Series | None:
        """Obtém a última linha de dados processados antes de uma data específica, permitindo consultas históricas.

        Args:
            before_date (datetime): Data antes da qual a última linha de dados processados deve ser obtida.

        Returns:
            pd.Series | None: A última linha de dados processados como uma Series do pandas, ou None se não houver dados antes da data especificada.
            """
        if not self.filepath.exists():
            return None

        df = self.load_sheet()
        if df.empty:
            return None

        df["date"] = pd.to_datetime(df["date"])
        df = df[df["date"] < before_date]
        if df.empty:
            return None

        return df.sort_values("date").iloc[-1]

    def get_data(self, year: int | None = None, month: int | None = None, filepath: Path | None = None) -> pd.DataFrame:
        """Obtém os dados processados para um período específico (ano e mês) ou a partir de um arquivo específico, permitindo flexibilidade na consulta dos dados.

            Args:
                year (int | None, optional): Ano para o qual os dados processados devem ser obtidos. Se None, não filtra por ano. Defaults to None.
                month (int | None, optional): Mês para o qual os dados processados devem ser obtidos. Se None, não filtra por mês. Defaults to None.
                filepath (Path | None, optional): Caminho para um arquivo específico de onde os dados processados devem ser obtidos. Se None, obtém os dados do armazenamento padrão. Defaults to None.

            Returns:
                pd.DataFrame: DataFrame contendo os dados processados para o período ou arquivo especificado
        """
        target_filepath = filepath if filepath else self.filepath

        if not target_filepath.exists():
            return self.schema_func()

        df = pd.read_excel(target_filepath)
        df["date"] = pd.to_datetime(df["date"])

        if year is not None:
            df = df[df["date"].dt.year == year]
        if month is not None:
            df = df[df["date"].dt.month == month]

        return df

    def order_by_date(self, df: pd.DataFrame) -> pd.DataFrame:
        """Ordena um DataFrame de dados processados por data, garantindo que os dados estejam organizados cronologicamente. Deve ser implementado por cada classe de armazenamento específica.

        Args:
            df (pd.DataFrame): DataFrame contendo os dados processados a serem ordenados.

        Returns:
            pd.DataFrame: DataFrame ordenado por data.
        """
        df["date"] = pd.to_datetime(df["date"])
        return df.sort_values("date")